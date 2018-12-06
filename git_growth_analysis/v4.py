import matplotlib
matplotlib.use("TkAgg")
import numpy as np 
import os
import scipy
import skimage
import operator
import tkinter as tk
import openpyxl

from numpy import array
from skimage import io 
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import NavigationToolbar2TkAgg
from statistics import mean
from scipy import optimize
from scipy.optimize import leastsq, fmin
from openpyxl import Workbook
from openpyxl import drawing

'''
Capture the interest points that we need
input: the image and the minimum intensity value
output: the image with only dots we need
'''
def capture(semi, m_val):
        green = semi[:,:,1]             #filter out all colors except green
        mask = green > m_val
        interest = green * mask         #apply mask
        return interest

def get_whole_data(mask):
    #put all .txt files in files
    files = []              #contains all of the .tiff files
    for f in os.listdir():
        if f.endswith(".tiff"):
            files.append(f)     #put every image in files

    #put all images intensity in a list, y axis
    intens_y = []
    for f in files:
        semi = io.imread(f)
        interest = capture(semi, mask)  #imply mask on each images                                      #this requires user input
        intens_y.append(interest.mean() * 100)  #put the average intensity in a list
    
    time_x = []
    for f in files:
        time_x.append(float(f[36] + f[37] + f[38] + f[39]))
    
    return files, time_x, intens_y

'''
Evaluate the user input mask by displaying the collected dots on the image with max and min
input: all available image intensity
output: none
'''
def evaluate_mask(files, intens_y, mask):
        print(intens_y)
        max_intens = max(intens_y)
        max_index = intens_y.index(max_intens)
        semi = io.imread(files[max_index])
        max_pic = capture(semi, mask)           #gets the max pic and filter it with our current MASK

        med_intens = np.median(intens_y)
        med_intens = intens_y.index(med_intens)
        semi = io.imread(files[med_intens])
        med_pic = capture(semi, mask)

        min_intens = min(intens_y)
        min_index = intens_y.index(min_intens)  #obtain min index
        semi = io.imread(files[min_index])
        min_pic = capture(semi, mask)
        
        plt.subplot(231).imshow(max_pic)        
        plt.subplot(232).imshow(med_pic)
        plt.subplot(233).imshow(min_pic)

def save_whole_excel(filename, time_x, intens_y):
        # print(filename)
        diction = dict(zip(time_x, intens_y))
        x_plot = []
        y_plot = []
        for key in sorted(diction.keys()):                                                                    
                x_plot.append(key)
                y_plot.append(diction[key])

        wb = Workbook()
        ws = wb.active

        ws.append(["Frames", "Time (sec)", "Intensity"])      #the title f the excel
        for i in range(len(x_plot)):
                ws.cell(row = i+2, column = 1).value = x_plot[i]
                ws.cell(row = i+2, column = 2).value = x_plot[i] * 0.2
                ws.cell(row = i+2, column = 3).value = y_plot[i]

        # plt.plot(x_plot, y_plot, 'ro')
        # plt.savefig("myplot.png", dpi = 150)
        # img = drawing.Image('myplot.png')
        # img.anchor(ws.cell('F1'))
        # ws.add_image(img)


        
        wb.save(filename + "_Time_Intensity_Plot.xlsx")

'''
obtain the max 5% of the data and min 5% of the data to perform scipy fit on it
input: mean, standard deviation, and x, y data
output: numpy arrays of x coordinates and y coordinates
'''
def maximum_minimum_scipy(m, std, dictionary):
        lowest = m - 1.3*std        #this only gives us the last 5% of the intensity
        lowest_cap = m - 3*std
        low_x = []
        low_y = []
        for key in sorted(dictionary.keys()):
                if(dictionary[key] < lowest and dictionary[key] > lowest_cap):
                        low_x.append(key)
                        low_y.append(dictionary[key])
        # plt.subplot(211)
        # plt.plot(low_x, low_y, 'ro')

        highest = m + 1.3*std        #this only gives us the top 5% of the intensity
        highest_cap = m + 3*std
        high_x = []
        high_y = []
        for key in sorted(dictionary.keys()):
                if(dictionary[key] > highest and dictionary[key] < highest_cap):
                        high_x.append(key)
                        high_y.append(dictionary[key])
        # plt.plot(high_x, high_y, 'bo')
        # plt.show()
        return (array(low_x+high_x), array(low_y+high_y))

def button0():
        path = os.getcwd() + "/" + str(entry0.get())    #get the path
        try:                                                    #if the file path exsits
                os.chdir(path)
                text_display = tk.Text(master=window, height = 1, width=18)
                text_display.grid(column=9, row=2)
                text_display.insert(tk.END, "Recorded!")
        except:                                                 #if there is no folder
                text_display = tk.Text(master=window, height = 1, width=18)
                text_display.grid(column=9, row=2)
                text_display.insert(tk.END, "No such file name!")

def button1():
        print(var.get())        #this is for the check box
        mask = int(entry1.get())

        files, time_x, intens_y = get_whole_data(mask)
        # evaluate_mask(files, intens_y, mask)
        # plt.subplot(212)
        filename = entry0.get()
        plt.plot(time_x, intens_y, 'ro')             #check whole graph (sinc)
        plt.show()
        if(var.get()):
                print(filename)
                save_whole_excel(filename, time_x, intens_y)


def button2():
        mask = int(entry1.get())
        start = int(entry2.get())
        guess_per = int(entry3.get())
        files, time_x, intens_y = get_whole_data(mask)

        #----------convert into dictionary to only get the sin part of the function
        diction = dict(zip(time_x, intens_y))

        #sort the dic by time and only capture the parts that we are interested, aka when we start growth
        x_plot = []
        y_plot = []
        for key in sorted(diction.keys()):
                if(key > start):                                                                     
                        x_plot.append(key)
                        y_plot.append(diction[key])
        x_np = array(x_plot)
        y_np = array(y_plot)
        dictionary = dict(zip(x_np, y_np))

        the_mean = np.mean(y_np)
        the_std = np.std(y_np)
        guess_phase = 0
        guess_period = guess_per                                                                              
        guess_freq = 1/guess_period
        guess_amp = the_std*2

        #perform algorithm fit!
        # optimize_func = lambda x: x[0]*np.sin(x[1]*x_np+x[2]) + x[3] - y_np
        optimize_func = lambda x: x[0]*np.sin(x[1]*x_np+x[2]) + the_mean
        est_amp, est_freq, est_phase, est_mean = leastsq(optimize_func, [guess_amp, guess_freq, guess_phase, the_mean])[0]
        fine_t = np.arange(start, max(x_np),0.1)
        print("Scipy: ", 1/est_freq)
        data_fit = est_amp*np.sin((fine_t*est_freq+est_phase)*2*np.pi) + the_mean     #equations for plot


        #take max std and min std part of the graph and scipy it
        mms_x, mms_y = maximum_minimum_scipy(the_mean, the_std, dictionary)
        mms_func = lambda x: x[0]*np.sin(x[1]*mms_x+x[2]) + the_mean
        mms_amp, mms_freq, mms_phase, mms_mean = leastsq(mms_func, [guess_amp, guess_freq, guess_phase, the_mean])[0]
        print("Max Min Scipy: ", 1/mms_freq)
        data_mms = guess_amp*np.sin((fine_t*mms_freq+mms_phase)*2*np.pi) + the_mean     #equations for plot

        plt.subplot(211)                   #for estimate
        plt.title('Scipy')
        plt.plot(x_plot, y_plot, 'ro')
        plt.plot(fine_t, data_fit)

        plt.subplot(212)                   #for max min scy, SOLO
        plt.title('Max Min Scipy')
        plt.plot(x_plot, y_plot, 'ro')
        plt.plot(fine_t, data_mms)
        plt.show()



#output GUI to obtain FILE name
window = tk.Tk()
window.title("GaN Growth Intensity Ploter")
window.geometry("550x250")
var = tk.IntVar()
r = 1
#gui for file name
label0 = tk.Label(text = "1. Enter File Name:").grid(column = 0, row = r)             
entry0 = tk.Entry(width = 13)
entry0.grid(column = 3, row = r)
button0 = tk.Button(text = "Record", command = button0).grid(column = 7, row = r+1)

#evaluate mask
label1 = tk.Label(text = "2. Enter Mask Value:").grid(column = 0, row = r+2)
entry1 = tk.Entry(width = 13)
entry1.grid(column = 3, row = r+2)
entry1.insert(0, "150")
check1 = tk.Checkbutton(window, text='save as .xlsx', variable=var).grid(column = 9, row = r+3)
button1 = tk.Button(text = "Record", command = button1).grid(column = 7, row = r+3)

#gui for guess starting point
label2 = tk.Label(text = "2. Enter Guess Starting Point:").grid(column = 0, row = r+4)
entry2 = tk.Entry(width = 13)
entry2.grid(column = 3, row = r+4)
#gui for guess period
label3 = tk.Label(text = "Guess Period:").grid(column = 0, row = r+5)
entry3 = tk.Entry(width = 13)
entry3.grid(column = 3, row = r+5)
button2 = tk.Button(text = "Record", command = button2).grid(column = 7, row = r+6)
window.mainloop()